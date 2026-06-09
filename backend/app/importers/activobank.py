from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any

from openpyxl import load_workbook

from app.importers.base import NormalisedTransaction


class ActivoBankImporter:
    source = "activobank"

    def parse_excel(self, file_content: bytes) -> list[NormalisedTransaction]:
        workbook = load_workbook(
            BytesIO(file_content),
            read_only=True,
            data_only=True,
        )

        sheet = workbook[workbook.sheetnames[0]]
        currency = self._find_currency(sheet)
        header_row_number = self._find_header_row(sheet)

        transactions: list[NormalisedTransaction] = []

        for row in sheet.iter_rows(
            min_row=header_row_number + 1,
            values_only=True,
        ):
            if self._is_empty_row(row):
                continue

            transaction = self._parse_row(row=row, currency=currency)
            transactions.append(transaction)

        return transactions

    def _find_currency(self, sheet) -> str:
        for row in sheet.iter_rows(min_row=1, max_row=10, values_only=True):
            if row[0] == "Moeda:" and row[1] is not None:
                return str(row[1]).strip().upper()

        return "EUR"

    def _find_header_row(self, sheet) -> int:
        for row_number, row in enumerate(
            sheet.iter_rows(min_row=1, max_row=20, values_only=True),
            start=1,
        ):
            values = [str(value).strip() if value is not None else "" for value in row]

            if values[:5] == ["Data Lanc.", "Data Valor", "Descrição", "Valor", "Saldo"]:
                return row_number

        raise ValueError("Could not find ActivoBank header row")

    def _parse_row(
        self,
        row: tuple[Any, ...],
        currency: str,
    ) -> NormalisedTransaction:
        transaction_date = self._parse_date(row[0])
        description = self._parse_description(row[2])
        amount = self._parse_amount(row[3])
        direction = "in" if amount > 0 else "out"

        return NormalisedTransaction(
            date=transaction_date,
            raw_description=description,
            description=description,
            amount=abs(amount),
            direction=direction,
            source=self.source,
            account="ActivoBank",
            currency=currency,
            external_id=None,
            notes=None,
        )

    def _parse_date(self, value: Any) -> date:
        if isinstance(value, datetime):
            return value.date()

        if isinstance(value, date):
            return value

        if isinstance(value, str):
            for date_format in ("%Y-%m-%d", "%d/%m/%Y"):
                try:
                    return datetime.strptime(value.strip(), date_format).date()
                except ValueError:
                    continue

        raise ValueError(f"Invalid ActivoBank date: {value}")

    def _parse_description(self, value: Any) -> str:
        if value is None:
            raise ValueError("Missing ActivoBank description")

        description = str(value).strip()

        if description == "":
            raise ValueError("Missing ActivoBank description")

        return description

    def _parse_amount(self, value: Any) -> Decimal:
        if value is None:
            raise ValueError("Missing ActivoBank amount")

        clean_value = str(value).strip().replace(",", ".")

        try:
            return Decimal(clean_value)
        except InvalidOperation as error:
            raise ValueError(f"Invalid ActivoBank amount: {value}") from error

    def _is_empty_row(self, row: tuple[Any, ...]) -> bool:
        return all(value is None or str(value).strip() == "" for value in row)
