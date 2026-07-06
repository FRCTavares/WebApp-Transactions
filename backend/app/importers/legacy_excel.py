from dataclasses import dataclass
from calendar import monthrange
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
import re
import unicodedata

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


SOURCE = "legacy_excel"
ACCOUNT = "manual_history"
CURRENCY = "EUR"

SKIPPED_SHEETS = {
    "painel central",
    "investments",
}

LAST_LEGACY_MONTH = date(2026, 4, 1)
LAST_LEGACY_WEALTH_MONTH = date(2026, 5, 31)

MONTHS_PT = {
    "janeiro": 1,
    "fevereiro": 2,
    "marco": 3,
    "março": 3,
    "abril": 4,
    "maio": 5,
    "junho": 6,
    "julho": 7,
    "agosto": 8,
    "setembro": 9,
    "outubro": 10,
    "novembro": 11,
    "dezembro": 12,
}


@dataclass(frozen=True)
class LegacyExcelTransaction:
    sheet_name: str
    row_number: int
    date: date
    description: str
    raw_description: str
    amount: Decimal
    direction: str
    cashflow_type: str
    source: str
    account: str
    currency: str
    category: str | None
    external_id: str
    notes: str | None


@dataclass(frozen=True)
class LegacyExcelOwedItem:
    sheet_name: str
    row_number: int
    person: str
    amount_total: Decimal
    amount_paid: Decimal
    reason: str
    status: str
    due_date: date | None
    notes: str | None
    external_id: str


@dataclass(frozen=True)
class LegacyExcelWealthSnapshot:
    sheet_name: str
    row_number: int
    column_number: int
    snapshot_date: date
    account_name: str
    account_type: str
    balance: Decimal
    currency: str
    balance_eur: Decimal
    fx_rate_to_eur: Decimal
    interest_earned: Decimal
    notes: str | None
    external_id: str


@dataclass(frozen=True)
class LegacyExcelInvalidRow:
    sheet_name: str
    row_number: int
    section: str
    error: str


@dataclass(frozen=True)
class LegacyExcelParseResult:
    transactions: list[LegacyExcelTransaction]
    owed_items: list[LegacyExcelOwedItem]
    invalid_rows: list[LegacyExcelInvalidRow]
    wealth_snapshots: list[LegacyExcelWealthSnapshot]


@dataclass(frozen=True)
class SectionLayout:
    name: str
    start_column: int
    end_column: int
    header_row: int
    columns: dict[str, int]


class LegacyExcelImporter:
    def parse_excel(self, file_content: bytes) -> LegacyExcelParseResult:
        workbook = load_workbook(BytesIO(file_content), data_only=True)

        transactions: list[LegacyExcelTransaction] = []
        owed_items: list[LegacyExcelOwedItem] = []
        invalid_rows: list[LegacyExcelInvalidRow] = []

        for sheet_name in workbook.sheetnames:
            if self._should_skip_sheet(sheet_name):
                continue

            sheet_month = self._parse_sheet_month(sheet_name)
            if sheet_month is None or sheet_month > LAST_LEGACY_MONTH:
                continue

            worksheet = workbook[sheet_name]
            layouts = self._find_section_layouts(worksheet)

            for layout in layouts:
                if layout.name == "expenses":
                    parsed, invalid = self._parse_expenses(
                        worksheet=worksheet,
                        layout=layout,
                        sheet_month=sheet_month,
                    )
                    transactions.extend(parsed)
                    invalid_rows.extend(invalid)

                if layout.name == "income":
                    parsed, invalid = self._parse_income(
                        worksheet=worksheet,
                        layout=layout,
                        sheet_month=sheet_month,
                    )
                    transactions.extend(parsed)
                    invalid_rows.extend(invalid)

                if layout.name == "owed":
                    parsed, invalid = self._parse_owed_items(
                        worksheet=worksheet,
                        layout=layout,
                        sheet_month=sheet_month,
                    )
                    owed_items.extend(parsed)
                    invalid_rows.extend(invalid)

        return LegacyExcelParseResult(
            transactions=transactions,
            owed_items=owed_items,
            invalid_rows=invalid_rows,
            wealth_snapshots=self.parse_wealth_snapshots(file_content),
        )


    def parse_wealth_snapshots(self, file_content: bytes) -> list[LegacyExcelWealthSnapshot]:
        workbook = load_workbook(BytesIO(file_content), data_only=True)
        worksheet = self._get_painel_central_sheet(workbook)

        if worksheet is None:
            return []

        month_columns = self._find_wealth_month_columns(worksheet)
        account_rows = self._find_wealth_account_rows(worksheet)

        snapshots: list[LegacyExcelWealthSnapshot] = []

        for account_key, account_config in account_rows.items():
            row_number = account_config["row_number"]

            for column_number, snapshot_date in month_columns.items():
                if snapshot_date > LAST_LEGACY_WEALTH_MONTH:
                    continue

                value = worksheet.cell(row_number, column_number).value

                if value is None:
                    continue

                try:
                    balance = self._parse_amount(value)
                except ValueError:
                    continue

                if balance <= 0:
                    continue

                account_name = account_config["account_name"]

                snapshots.append(
                    LegacyExcelWealthSnapshot(
                        sheet_name=worksheet.title,
                        row_number=row_number,
                        column_number=column_number,
                        snapshot_date=snapshot_date,
                        account_name=account_name,
                        account_type=account_config["account_type"],
                        balance=balance,
                        currency=CURRENCY,
                        balance_eur=balance,
                        fx_rate_to_eur=Decimal("1"),
                        interest_earned=Decimal("0.00"),
                        notes="Legacy Excel wealth snapshot.",
                        external_id=self._build_wealth_external_id(
                            account_key=account_key,
                            snapshot_date=snapshot_date,
                        ),
                    )
                )

        return snapshots

    def _get_painel_central_sheet(self, workbook):
        for sheet_name in workbook.sheetnames:
            if self._normalise_text(sheet_name) == "painel central":
                return workbook[sheet_name]

        return None

    def _find_wealth_month_columns(self, worksheet: Worksheet) -> dict[int, date]:
        month_columns: dict[int, date] = {}

        for row in worksheet.iter_rows():
            for cell in row:
                snapshot_date = self._parse_wealth_month_header(cell.value)

                if snapshot_date is not None:
                    month_columns[cell.column] = snapshot_date

        return month_columns

    def _parse_wealth_month_header(self, value) -> date | None:
        if value is None:
            return None

        text = self._normalise_text(value)
        parts = text.split()

        if len(parts) < 2:
            return None

        month = MONTHS_PT.get(parts[0])
        if month is None:
            return None

        year_text = parts[-1]
        if not year_text.isdigit():
            return None

        year = int(year_text)
        if year < 100:
            year += 2000

        last_day = monthrange(year, month)[1]
        return date(year, month, last_day)

    def _find_wealth_account_rows(self, worksheet: Worksheet) -> dict[str, dict[str, int | str]]:
        account_configs = {
            "activobank": {
                "matches": {"activobank", "activo bank"},
                "account_name": "ActivoBank",
                "account_type": "current_account",
            },
            "revolut": {
                "matches": {"revolut"},
                "account_name": "Revolut",
                "account_type": "current_account",
            },
            "trading_212": {
                "matches": {"trading 212", "trading212"},
                "account_name": "Trading 212",
                "account_type": "brokerage",
            },
            "bank_notes": {
                "matches": {"bank notes", "cash", "notas"},
                "account_name": "Bank Notes",
                "account_type": "cash",
            },
            "money_owed_to_me": {
                "matches": {"dividas nao pagas", "dívidas não pagas"},
                "account_name": "Money Owed To Me",
                "account_type": "other",
            },
        }

        rows: dict[str, dict[str, int | str]] = {}

        for row in worksheet.iter_rows():
            for cell in row:
                value = self._normalise_text(cell.value)

                if not value:
                    continue

                for account_key, config in account_configs.items():
                    if value in config["matches"]:
                        rows[account_key] = {
                            "row_number": cell.row,
                            "account_name": config["account_name"],
                            "account_type": config["account_type"],
                        }

        return rows

    def _build_wealth_external_id(self, account_key: str, snapshot_date: date) -> str:
        return f"{SOURCE}:wealth:{account_key}:{snapshot_date.isoformat()}"


    def _should_skip_sheet(self, sheet_name: str) -> bool:
        return self._normalise_text(sheet_name) in SKIPPED_SHEETS

    def _parse_sheet_month(self, sheet_name: str) -> date | None:
        parts = sheet_name.strip().split()
        if len(parts) < 2:
            return None

        month_name = self._normalise_text(parts[0])
        month = MONTHS_PT.get(month_name)
        if month is None:
            return None

        year_text = parts[-1]
        if not year_text.isdigit():
            return None

        year = int(year_text)
        if year < 100:
            year += 2000

        return date(year, month, 1)

    def _find_section_layouts(self, worksheet: Worksheet) -> list[SectionLayout]:
        section_cells: list[tuple[str, int, int]] = []

        for row in worksheet.iter_rows():
            for cell in row:
                value = self._normalise_text(cell.value)
                if not value:
                    continue

                if "despesas" in value:
                    section_cells.append(("expenses", cell.row, cell.column))
                elif "rendimentos" in value:
                    section_cells.append(("income", cell.row, cell.column))
                elif "dividas de outros" in value:
                    section_cells.append(("owed", cell.row, cell.column))

        if not section_cells:
            return []

        section_cells = sorted(
            section_cells,
            key=lambda section_cell: (section_cell[1], section_cell[2]),
        )

        layouts: list[SectionLayout] = []

        for index, (name, row_number, start_column) in enumerate(section_cells):
            next_start_column = worksheet.max_column + 1

            for _next_name, next_row, next_column in section_cells[index + 1:]:
                if next_row == row_number:
                    next_start_column = next_column
                    break

            header_row = row_number + 1
            end_column = next_start_column - 1
            columns = self._map_header_columns(
                worksheet=worksheet,
                header_row=header_row,
                start_column=start_column,
                end_column=end_column,
            )

            layouts.append(
                SectionLayout(
                    name=name,
                    start_column=start_column,
                    end_column=end_column,
                    header_row=header_row,
                    columns=columns,
                )
            )

        return layouts

    def _map_header_columns(
        self,
        worksheet: Worksheet,
        header_row: int,
        start_column: int,
        end_column: int,
    ) -> dict[str, int]:
        columns: dict[str, int] = {}

        for column in range(start_column, end_column + 1):
            header = self._normalise_text(worksheet.cell(header_row, column).value)

            if header == "valor":
                columns["amount"] = column
            elif header == "tipo":
                columns["type"] = column
            elif header == "categoria":
                columns["category"] = column
            elif header == "descricao":
                columns["description"] = column
            elif header == "deve-me":
                columns["owed_by"] = column
            elif header == "quem":
                columns["person"] = column
            elif header == "pagou?":
                columns["paid"] = column

        return columns

    def _parse_expenses(
        self,
        worksheet: Worksheet,
        layout: SectionLayout,
        sheet_month: date,
    ) -> tuple[list[LegacyExcelTransaction], list[LegacyExcelInvalidRow]]:
        transactions: list[LegacyExcelTransaction] = []
        invalid_rows: list[LegacyExcelInvalidRow] = []

        for row_number in range(layout.header_row + 1, worksheet.max_row + 1):
            amount_value = self._get_cell_value(worksheet, row_number, layout.columns.get("amount"))
            description_value = self._get_cell_value(
                worksheet,
                row_number,
                layout.columns.get("description"),
            )

            if self._is_empty_row(amount_value, description_value):
                continue

            if description_value is None:
                continue

            try:
                amount = self._parse_amount(amount_value)
                if amount <= 0:
                    continue

                description = self._parse_description(description_value)
                category = self._parse_optional_text(
                    self._get_cell_value(
                        worksheet,
                        row_number,
                        layout.columns.get("category") or layout.columns.get("type"),
                    )
                )
                owed_by = self._parse_optional_text(
                    self._get_cell_value(worksheet, row_number, layout.columns.get("owed_by"))
                )

                cashflow_type = "expense"

                transactions.append(
                    LegacyExcelTransaction(
                        sheet_name=worksheet.title,
                        row_number=row_number,
                        date=self._infer_row_date(description_value, sheet_month),
                        description=description,
                        raw_description=self._build_raw_description(
                            worksheet.title,
                            "Despesas",
                            row_number,
                            description_value,
                        ),
                        amount=amount,
                        direction="out",
                        cashflow_type=cashflow_type,
                        source=SOURCE,
                        account=ACCOUNT,
                        currency=CURRENCY,
                        category=category,
                        external_id=self._build_external_id(
                            worksheet.title,
                            "expenses",
                            row_number,
                        ),
                        notes=f"Legacy Excel import. Owed by: {owed_by}" if owed_by else "Legacy Excel import.",
                    )
                )
            except ValueError as error:
                invalid_rows.append(
                    LegacyExcelInvalidRow(
                        sheet_name=worksheet.title,
                        row_number=row_number,
                        section="Despesas",
                        error=str(error),
                    )
                )

        return transactions, invalid_rows

    def _parse_income(
        self,
        worksheet: Worksheet,
        layout: SectionLayout,
        sheet_month: date,
    ) -> tuple[list[LegacyExcelTransaction], list[LegacyExcelInvalidRow]]:
        transactions: list[LegacyExcelTransaction] = []
        invalid_rows: list[LegacyExcelInvalidRow] = []

        for row_number in range(layout.header_row + 1, worksheet.max_row + 1):
            amount_value = self._get_cell_value(worksheet, row_number, layout.columns.get("amount"))
            description_value = self._get_cell_value(
                worksheet,
                row_number,
                layout.columns.get("description"),
            )

            if self._is_empty_row(amount_value, description_value):
                continue

            if description_value is None:
                continue

            paid_value = self._get_cell_value(worksheet, row_number, layout.columns.get("paid"))
            if layout.columns.get("paid") is not None and not self._is_paid(paid_value):
                continue

            try:
                amount = self._parse_amount(amount_value)
                if amount <= 0:
                    continue

                description = self._parse_description(description_value)
                category = self._parse_optional_text(
                    self._get_cell_value(
                        worksheet,
                        row_number,
                        layout.columns.get("category") or layout.columns.get("type"),
                    )
                )

                transactions.append(
                    LegacyExcelTransaction(
                        sheet_name=worksheet.title,
                        row_number=row_number,
                        date=self._infer_row_date(description_value, sheet_month),
                        description=description,
                        raw_description=self._build_raw_description(
                            worksheet.title,
                            "Rendimentos",
                            row_number,
                            description_value,
                        ),
                        amount=amount,
                        direction="in",
                        cashflow_type="income",
                        source=SOURCE,
                        account=ACCOUNT,
                        currency=CURRENCY,
                        category=category,
                        external_id=self._build_external_id(
                            worksheet.title,
                            "income",
                            row_number,
                        ),
                        notes="Legacy Excel import.",
                    )
                )
            except ValueError as error:
                invalid_rows.append(
                    LegacyExcelInvalidRow(
                        sheet_name=worksheet.title,
                        row_number=row_number,
                        section="Rendimentos",
                        error=str(error),
                    )
                )

        return transactions, invalid_rows

    def _parse_owed_items(
        self,
        worksheet: Worksheet,
        layout: SectionLayout,
        sheet_month: date,
    ) -> tuple[list[LegacyExcelOwedItem], list[LegacyExcelInvalidRow]]:
        owed_items: list[LegacyExcelOwedItem] = []
        invalid_rows: list[LegacyExcelInvalidRow] = []

        for row_number in range(layout.header_row + 1, worksheet.max_row + 1):
            amount_value = self._get_cell_value(worksheet, row_number, layout.columns.get("amount"))
            description_value = self._get_cell_value(
                worksheet,
                row_number,
                layout.columns.get("description"),
            )

            if self._is_empty_row(amount_value, description_value):
                continue

            if description_value is None:
                continue

            try:
                amount = self._parse_amount(amount_value)
                if amount <= 0:
                    continue

                reason = self._parse_description(description_value)
                paid = self._is_paid(
                    self._get_cell_value(worksheet, row_number, layout.columns.get("paid"))
                )
                person = self._parse_optional_text(
                    self._get_cell_value(worksheet, row_number, layout.columns.get("person"))
                )

                owed_items.append(
                    LegacyExcelOwedItem(
                        sheet_name=worksheet.title,
                        row_number=row_number,
                        person=person or "Unknown",
                        amount_total=amount,
                        amount_paid=amount if paid else Decimal("0.00"),
                        reason=reason,
                        status="paid" if paid else "open",
                        due_date=self._infer_row_date(description_value, sheet_month),
                        notes="Legacy Excel import.",
                        external_id=self._build_external_id(
                            worksheet.title,
                            "owed",
                            row_number,
                        ),
                    )
                )
            except ValueError as error:
                invalid_rows.append(
                    LegacyExcelInvalidRow(
                        sheet_name=worksheet.title,
                        row_number=row_number,
                        section="Dívidas de Outros",
                        error=str(error),
                    )
                )

        return owed_items, invalid_rows

    def _get_cell_value(self, worksheet: Worksheet, row_number: int, column: int | None):
        if column is None:
            return None

        return worksheet.cell(row_number, column).value

    def _is_empty_row(self, amount_value, description_value) -> bool:
        return amount_value is None and description_value is None

    def _parse_amount(self, value) -> Decimal:
        if value is None:
            raise ValueError("Missing amount")

        if isinstance(value, Decimal):
            amount = value
        elif isinstance(value, int | float):
            amount = Decimal(str(value))
        else:
            cleaned = str(value).strip().replace("€", "").replace(" ", "").replace(",", ".")
            try:
                amount = Decimal(cleaned)
            except InvalidOperation as error:
                raise ValueError(f"Invalid amount: {value}") from error

        return amount.quantize(Decimal("0.01"))

    def _parse_description(self, value) -> str:
        if value is None:
            raise ValueError("Missing description")

        if isinstance(value, datetime):
            return value.date().isoformat()

        if isinstance(value, date):
            return value.isoformat()

        description = str(value).strip()
        if not description:
            raise ValueError("Missing description")

        return description

    def _parse_optional_text(self, value) -> str | None:
        if value is None:
            return None

        text = str(value).strip()
        if not text:
            return None

        return text

    def _infer_row_date(self, description_value, sheet_month: date) -> date:
        if isinstance(description_value, datetime):
            return description_value.date()

        if isinstance(description_value, date):
            return description_value

        description = "" if description_value is None else str(description_value)
        date_match = re.search(r"\b(\d{1,2})[/-](\d{1,2})\b", description)

        if not date_match:
            return sheet_month

        day = int(date_match.group(1))
        month = int(date_match.group(2))

        if month != sheet_month.month:
            return sheet_month

        try:
            return date(sheet_month.year, month, day)
        except ValueError:
            return sheet_month

    def _is_paid(self, value) -> bool:
        text = self._normalise_text(value)
        return text in {"sim", "s", "yes", "y", "true", "pago", "paid"}

    def _build_raw_description(
        self,
        sheet_name: str,
        section: str,
        row_number: int,
        description_value,
    ) -> str:
        return f"{sheet_name} | {section} | row {row_number} | {description_value}"

    def _build_external_id(
        self,
        sheet_name: str,
        section: str,
        row_number: int,
    ) -> str:
        return f"{SOURCE}:{sheet_name}:{section}:{row_number}"

    def _normalise_text(self, value) -> str:
        if value is None:
            return ""

        text = str(value).strip().lower()
        text = unicodedata.normalize("NFKD", text)
        text = "".join(character for character in text if not unicodedata.combining(character))

        return " ".join(text.split())
