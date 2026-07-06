from io import BytesIO
from decimal import Decimal

from openpyxl import Workbook

from app.importers.legacy_excel import LegacyExcelImporter


def build_workbook_bytes() -> bytes:
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    investments = workbook.create_sheet("Investments")
    investments["A1"] = "Should be skipped"
    investments["B2"] = "Despesas"
    investments["B3"] = "Valor"
    investments["C3"] = "Descrição"
    investments["B4"] = 999
    investments["C4"] = "Investment row"

    old_sheet = workbook.create_sheet("Setembro 24")
    old_sheet["B2"] = "Despesas Setembro de 2024"
    old_sheet["G2"] = "Rendimentos Setembro de 2024"
    old_sheet["K2"] = "Dívidas de Outros"
    old_sheet["B3"] = "Valor"
    old_sheet["C3"] = "Tipo"
    old_sheet["D3"] = "Descrição"
    old_sheet["E3"] = "Deve-me"
    old_sheet["G3"] = "Valor"
    old_sheet["H3"] = "Tipo"
    old_sheet["I3"] = "Descrição"
    old_sheet["K3"] = "Valor"
    old_sheet["L3"] = "Descrição"
    old_sheet["M3"] = "Quem"
    old_sheet["N3"] = "Pagou?"

    old_sheet["B4"] = 27.89
    old_sheet["C4"] = "Saúde"
    old_sheet["D4"] = "Wells teste"
    old_sheet["E4"] = "Martinha"
    old_sheet["G4"] = 130
    old_sheet["H4"] = "Mesada"
    old_sheet["I4"] = "Pai pagou a mesada de Setembro"
    old_sheet["K4"] = 8.53
    old_sheet["L4"] = "Pagou metade das compras"
    old_sheet["M4"] = "Martinha"
    old_sheet["N4"] = "Sim"

    mixed_sheet = workbook.create_sheet("Junho 25")
    mixed_sheet["B2"] = "Despesas"
    mixed_sheet["G2"] = "Rendimentos "
    mixed_sheet["L2"] = "Dívidas de Outros"
    mixed_sheet["B3"] = "Valor"
    mixed_sheet["C3"] = "Tipo"
    mixed_sheet["D3"] = "Descrição"
    mixed_sheet["E3"] = "Deve-me"
    mixed_sheet["G3"] = "Valor"
    mixed_sheet["H3"] = "Tipo"
    mixed_sheet["I3"] = "Descrição"
    mixed_sheet["J3"] = "Pagou?"
    mixed_sheet["L3"] = "Valor"
    mixed_sheet["M3"] = "Descrição"
    mixed_sheet["N3"] = "Quem"
    mixed_sheet["O3"] = "Pagou?"

    mixed_sheet["B4"] = 3.98
    mixed_sheet["C4"] = "Super"
    mixed_sheet["D4"] = "04/06 Minipreço"
    mixed_sheet["E4"] = "Mãe"
    mixed_sheet["G4"] = 255
    mixed_sheet["H4"] = "Mensal"
    mixed_sheet["I4"] = "Mesada + Salário Pai"
    mixed_sheet["J4"] = "sIM"
    mixed_sheet["G5"] = 80
    mixed_sheet["H5"] = "Mensal"
    mixed_sheet["I5"] = "Mesada não recebida"
    mixed_sheet["J5"] = "Não"
    mixed_sheet["L4"] = 6.74
    mixed_sheet["M4"] = "04/06 Minipreço"
    mixed_sheet["N4"] = "Mãe"
    mixed_sheet["O4"] = "Não"

    new_sheet = workbook.create_sheet("Fevereiro 26")
    new_sheet["B2"] = "Despesas"
    new_sheet["I2"] = "Rendimentos "
    new_sheet["M2"] = "Dívidas de Outros"
    new_sheet["B3"] = "Valor"
    new_sheet["C3"] = "Descrição"
    new_sheet["D3"] = "Deve-me"
    new_sheet["E3"] = "Categoria"
    new_sheet["I3"] = "Valor"
    new_sheet["J3"] = "Descrição"
    new_sheet["K3"] = "Pagou?"
    new_sheet["M3"] = "Valor"
    new_sheet["N3"] = "Descrição"
    new_sheet["O3"] = "Quem"
    new_sheet["P3"] = "Pagou?"

    new_sheet["B4"] = 31.9
    new_sheet["C4"] = "Ginásio Fevereiro"
    new_sheet["E4"] = "Fixo"
    new_sheet["I4"] = 280
    new_sheet["J4"] = "Mesada + Salario Pai"
    new_sheet["K4"] = "Sim"
    new_sheet["M4"] = 14.66
    new_sheet["N4"] = "Chat parte do gui e marta"
    new_sheet["P4"] = "Sim"

    skipped_new_sheet = workbook.create_sheet("Maio 26")
    skipped_new_sheet["B2"] = "Despesas"
    skipped_new_sheet["B3"] = "Valor"
    skipped_new_sheet["C3"] = "Descrição"
    skipped_new_sheet["B4"] = 999
    skipped_new_sheet["C4"] = "Should be skipped because app era starts May 2026"

    output = BytesIO()
    workbook.save(output)

    return output.getvalue()


def test_legacy_excel_importer_skips_investments_and_may_2026():
    result = LegacyExcelImporter().parse_excel(build_workbook_bytes())

    sheet_names = {
        item.sheet_name
        for item in [*result.transactions, *result.owed_items]
    }

    assert "Investments" not in sheet_names
    assert "Maio 26" not in sheet_names


def test_legacy_excel_importer_parses_transactions_and_owed_items():
    result = LegacyExcelImporter().parse_excel(build_workbook_bytes())

    assert len(result.invalid_rows) == 0
    assert len(result.transactions) == 6
    assert len(result.owed_items) == 3

    expense = next(
        transaction
        for transaction in result.transactions
        if transaction.description == "04/06 Minipreço"
    )
    assert expense.date.isoformat() == "2025-06-04"
    assert expense.amount == Decimal("3.98")
    assert expense.direction == "out"
    assert expense.cashflow_type == "expense"
    assert expense.category == "Super"
    assert expense.source == "legacy_excel"
    assert expense.account == "manual_history"
    assert expense.currency == "EUR"

    income_descriptions = {
        transaction.description
        for transaction in result.transactions
        if transaction.direction == "in"
    }
    assert "Mesada não recebida" not in income_descriptions

    owed = next(
        owed_item
        for owed_item in result.owed_items
        if owed_item.reason == "Chat parte do gui e marta"
    )
    assert owed.amount_total == Decimal("14.66")
    assert owed.amount_paid == Decimal("14.66")
    assert owed.status == "paid"
    assert owed.person == "Unknown"


def test_legacy_excel_importer_handles_new_category_layout():
    result = LegacyExcelImporter().parse_excel(build_workbook_bytes())

    gym = next(
        transaction
        for transaction in result.transactions
        if transaction.description == "Ginásio Fevereiro"
    )

    assert gym.category == "Fixo"
    assert gym.cashflow_type == "expense"
    assert gym.external_id == "legacy_excel:Fevereiro 26:expenses:4"


def test_legacy_excel_importer_does_not_treat_fractions_as_dates():
    result = LegacyExcelImporter().parse_excel(build_workbook_bytes())

    owed_dates = {
        owed_item.reason: owed_item.due_date
        for owed_item in result.owed_items
    }

    assert owed_dates["Pagou metade das compras"].isoformat() == "2024-09-01"
